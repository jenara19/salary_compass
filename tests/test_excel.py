"""
Smoke tests for output/excel.py — generate_excel_report().

These tests verify that:
- The function returns non-empty bytes
- The bytes form a valid ZIP/XLSX file
- All 5 expected sheets are present in the workbook
- Sheet 2 (Budget Breakdown) correctly reads `value_eur` from budget items (B1 regression)
"""

import io
import zipfile

import openpyxl

from output.excel import generate_excel_report


def _minimal_inputs(with_budget_items: bool = False):
    """Minimal but structurally complete inputs for the export function."""
    city_inputs = {
        "amsterdam": {
            "country_code": "NL",
            "currency": "EUR",
            "gross_annual": 85000,
            "effective_gross": 85000,
            "active_schemes": ["nl_30_ruling"],
            "scheme_expiry": {"nl_30_ruling": 5},
            "ruling_start_year": 2026,
            "cagr": 0.03,
            "perks_monthly_eur": 100.0,
            "city_overrides": {},
            "bonus_pct": 0.0,
            "rsu": 0.0,
            "company": "",
        },
        "berlin": {
            "country_code": "DE",
            "currency": "EUR",
            "gross_annual": 90000,
            "effective_gross": 90000,
            "active_schemes": [],
            "scheme_expiry": {},
            "ruling_start_year": 2026,
            "cagr": 0.03,
            "perks_monthly_eur": 0.0,
            "city_overrides": {},
            "bonus_pct": 0.0,
            "rsu": 0.0,
            "company": "",
        },
    }

    scenarios_def = [
        {
            "name": "comfortable",
            "label": "Comfortable",
            "category_multipliers": {
                "rent": 1.0,
                "food": 1.0,
                "transport": 1.0,
                "healthcare": 1.0,
                "entertainment": 1.0,
                "misc": 1.0,
            },
        }
    ]

    city_names = {"amsterdam": "Amsterdam", "berlin": "Berlin"}
    scen_names = ["comfortable"]

    # Build a minimal results_matrix with correct structure
    from engine.tax import calculate_net
    from engine.trajectory import calculate_trajectory

    # Sample budget items using the correct key `value_eur`
    sample_items = {
        "rent_2bed": {"label": "Rent / Housing", "value_eur": 1500, "value_local": 1500},
        "food": {"label": "Food", "value_eur": 400, "value_local": 400},
        "transport": {"label": "Transport", "value_eur": 150, "value_local": 150},
    }

    results_matrix = {}
    for slug, ci in city_inputs.items():
        results_matrix[slug] = {}
        for scen in scenarios_def:
            net = calculate_net(ci["effective_gross"], ci["country_code"], ci["active_schemes"])
            traj = calculate_trajectory(
                ci["effective_gross"],
                ci["country_code"],
                slug,
                active_schemes=ci["active_schemes"],
                scheme_expiry=ci["scheme_expiry"],
                expenses_eur=2500,
                cagr=ci["cagr"],
                ruling_start_year=ci["ruling_start_year"],
                perks_monthly_eur=ci["perks_monthly_eur"],
            )
            budget = {
                "total_eur": 2050,
                "items": sample_items if with_budget_items else {},
            }
            results_matrix[slug][scen["name"]] = {
                "net": net,
                "trajectory": traj,
                "budget": budget,
            }

    return city_inputs, results_matrix, scenarios_def, city_names, scen_names


class TestExcelSmoke:
    def test_returns_bytes(self):
        city_inputs, results_matrix, scenarios_def, city_names, scen_names = _minimal_inputs()
        raw = generate_excel_report(
            city_inputs=city_inputs,
            results_matrix=results_matrix,
            scenarios_def=scenarios_def,
            city_names=city_names,
            home_city_slug="amsterdam",
            scen_names=scen_names,
        )
        assert isinstance(raw, bytes), "generate_excel_report must return bytes"
        assert len(raw) > 0, "Returned bytes must be non-empty"

    def test_valid_zip_structure(self):
        """XLSX is a ZIP file — verify it can be opened as such."""
        city_inputs, results_matrix, scenarios_def, city_names, scen_names = _minimal_inputs()
        raw = generate_excel_report(
            city_inputs=city_inputs,
            results_matrix=results_matrix,
            scenarios_def=scenarios_def,
            city_names=city_names,
            home_city_slug="amsterdam",
            scen_names=scen_names,
        )
        buf = io.BytesIO(raw)
        assert zipfile.is_zipfile(buf), "Output must be a valid ZIP/XLSX file"

    def test_five_sheets_present(self):
        """Verify the workbook contains all 5 expected sheets."""
        city_inputs, results_matrix, scenarios_def, city_names, scen_names = _minimal_inputs()
        raw = generate_excel_report(
            city_inputs=city_inputs,
            results_matrix=results_matrix,
            scenarios_def=scenarios_def,
            city_names=city_names,
            home_city_slug="amsterdam",
            scen_names=scen_names,
        )
        buf = io.BytesIO(raw)
        with zipfile.ZipFile(buf) as zf:
            names = zf.namelist()
            # xl/worksheets/sheet*.xml correspond to each sheet
            sheet_files = [n for n in names if n.startswith("xl/worksheets/sheet")]
        assert len(sheet_files) == 5, f"Expected 5 sheets, found {len(sheet_files)}: {sheet_files}"


class TestExcelBudgetBreakdown:
    """Regression tests for B1: budget items must use `value_eur` key (not `eur`)."""

    def test_budget_cells_non_zero_when_items_provided(self):
        """Budget breakdown sheet must show non-zero values when items have `value_eur`."""
        city_inputs, results_matrix, scenarios_def, city_names, scen_names = _minimal_inputs(
            with_budget_items=True
        )
        raw = generate_excel_report(
            city_inputs=city_inputs,
            results_matrix=results_matrix,
            scenarios_def=scenarios_def,
            city_names=city_names,
            home_city_slug="amsterdam",
            scen_names=scen_names,
        )
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        sheet_names = wb.sheetnames
        budget_sheet = next((s for s in sheet_names if "Budget" in s or "2." in s), None)
        assert budget_sheet is not None, f"Budget sheet not found in: {sheet_names}"
        ws = wb[budget_sheet]

        # Collect all numeric cell values from the sheet
        numeric_values = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, (int, float))
        ]
        # With our sample_items, total=2050, and individual items have values like 1500, 400, 150
        assert any(v > 0 for v in numeric_values), (
            f"Budget breakdown sheet has no positive numeric values — "
            f"likely reading wrong key ('eur' instead of 'value_eur'). "
            f"Found values: {numeric_values[:20]}"
        )

    def test_budget_cells_total_row_correct(self):
        """Total row in budget sheet should match sum of items (budget['total_eur'])."""
        city_inputs, results_matrix, scenarios_def, city_names, scen_names = _minimal_inputs(
            with_budget_items=True
        )
        raw = generate_excel_report(
            city_inputs=city_inputs,
            results_matrix=results_matrix,
            scenarios_def=scenarios_def,
            city_names=city_names,
            home_city_slug="amsterdam",
            scen_names=scen_names,
        )
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        sheet_names = wb.sheetnames
        budget_sheet = next((s for s in sheet_names if "Budget" in s or "2." in s), None)
        assert budget_sheet is not None
        ws = wb[budget_sheet]

        # Find the TOTAL row and check value for Amsterdam (column B = col index 2)
        for row in ws.iter_rows():
            cells = list(row)
            if cells and cells[0].value == "TOTAL":
                total_val = cells[1].value  # col B = Amsterdam
                assert total_val == 2050, f"TOTAL row should be 2050, got {total_val}"
                break
